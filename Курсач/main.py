"""
Программный продукт может быть:

    законченным самостоятельным приложением;

    модулем или пакетом расширения, подключаемым к поддерживающему 
    его приложению;

    процедурой или набором процедур (функций) в составе законченного 
    приложения;

    законченным программным кодом на интерпретируемом языке 
    программирования, поддерживаемом некоторым приложением.

    СУД должна работать с несколькими ролями: администратор, 
    пользователь, суперпользователь

    СУД должна обеспечивать многопользовательский режим

    СУД должна обеспечивать защиту на уровне строк (RLS) 

    СУД должна обеспечивать различный и в то же время полный клиентский 
    интерфейс в зависимости от роли.
"""



# Оператор скорой помощи (имеет доступ ко всем таблицам. Добавление,
# удаление заявок на вызов скорой помощи, добавление больных)

# Врач
# (имеет доступ к таблицам заявок, больные. Может только
# просматривать таблицу заявок, просматривать, удалять больных)

# Медсестра (имеет доступ только к таблице больные. Может только
# просматривать таблицу больные)


from typing import List, Any, Union, Dict
from PyQt5 import QtGui, QtWidgets, QtCore
from UI_Forms.main_formUI import Ui_MainWindow
import psycopg2

import sys

from numpy import round as np_round
import matplotlib.pyplot as plt

import openpyxl
import datetime

from CONSTANTS import COLUMNS_NAMES, REVERSED_COLUMNS_NAMES, TABLES_DICT, CHILDREN_TABLES, \
    MODIFIED_VIEW, QUERIES, PARAMS

from CompoundForm import CompoundForm


if "--test" in sys.argv:
    TESTING_ENABLED = 1
else:
    TESTING_ENABLED = 0

def showMessage(text: str) -> None:
    msg = QtWidgets.QMessageBox()
    msg.setIcon(QtWidgets.QMessageBox.Information)
    msg.setText(text)
    msg.setWindowTitle("Info")
    msg.exec_()

def showError(text: str) -> None:
    msg = QtWidgets.QMessageBox()
    msg.setIcon(QtWidgets.QMessageBox.Critical)
    msg.setText(text)
    msg.setWindowTitle("Info")
    msg.exec_()


def drawPieChart(values: List[int], labels: List[str], title: str) -> Union[str, None]:
    if not len(values) == len(labels):
        return "Значения и метки должны быть одинаковой длины"

    max_index_value = values.index(max(values))
    get_value = lambda x: np_round(x / 100 * sum(values))

    plt.pie(values, labels=labels, explode=tuple([0 for _ in values[:max_index_value]]) + tuple([.1]) + tuple([0 for _ in values[max_index_value + 1:]]), autopct=get_value)
    plt.title(title)
    plt.show()

def saveDataToExcel(columns_names: List[str], values: List[List[Any]], report_name: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active

    for idx, column in enumerate(columns_names):
        ws.cell(row=1, column=idx + 1).value = column
        ws.cell(row=1, column=idx + 1).font = openpyxl.styles.Font(bold=True)

    for row_idx, row in enumerate(values):
        for column_idx, column in enumerate(row):
            try:
                ws.cell(row=row_idx + 2, column=column_idx + 1).value = column
            except Exception as e:
                return showError(str(e))

    current_datetime = datetime.datetime.now().strftime("%d%m%y%H%M%S")
    wb_name = f"Отчёт по {report_name} {current_datetime}.xlsx"
    wb.save(wb_name)


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    window_closed = QtCore.pyqtSignal()
    def __init__(self, cursor, conn) -> None:
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self._cursor = cursor 
        self._conn = conn

        self._table_is_displayed = False

        self._new_rows_added: List[int] = list()
        self._updatedRecordsInfo: Dict[str, Dict[str, str]] = dict()
        self._default_procedures_names: List[str] = list()

        self.visualizeTable_pushButton.pressed.connect(self._visualizeTable)
        self.visualizeQuery_pushButton.pressed.connect(self._visualizeQuery)
        self.saveToExcel_pushButton.pressed.connect(self._saveToExcel)
        self.buildSummaryChart_pushButton.pressed.connect(self._buildSummaryChart)
        self.findByCriteria_pushButton.pressed.connect(self._findByCriteria)

        self.buildSummaryChart_pushButton.setEnabled(False)

        self.addRow_pushButton.pressed.connect(self._addRowToTheTableWidget)
        self.addRecord_pushButton.pressed.connect(self._addRecord)
        self.updateRecord_pushButton.pressed.connect(self._updateRecord)
        self.deleteRecord_pushButton.pressed.connect(self._deleteRecord)

        self.tables_comboBox.currentTextChanged.connect(self._tablesChangedEvent)
        self.queries_comboBox.currentTextChanged.connect(self._queriesChangedEvent)

        self.params_textEdit.textChanged.connect(self._checkParamsAmount)
        
        self.editChildTable_pushButton.pressed.connect(self._prepareCompoundForm)

        self.enableSorting_checkBox.stateChanged.connect(self._updateSortingButtonsState)

        self.centralWidget().setLayout(self.main_verticalLayout)

        self.__setComboBoxes()

    def __setComboBoxes(self) -> None:
        self.__setTables()
        self.__setQueries()

    def __setTables(self) -> None:
        for table in TABLES_DICT.keys():
            self.tables_comboBox.addItem(table)

    def __setQueries(self) -> None:
        for query in QUERIES.keys():
            self.queries_comboBox.addItem(query)

    def closeEvent(self, a0: QtGui.QCloseEvent) -> None:
        self.window_closed.emit()
        return super().closeEvent(a0)

    def __clearTableWidget(self) -> None:
        self.tableWidget.clear()
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(0)
        self.buildSummaryChart_pushButton.setEnabled(False)

    def _tablesChangedEvent(self) -> None:
        self.__fillColumns()
        self.__clearTableWidget()

    def _queriesChangedEvent(self) -> None:
        query_name = self.queries_comboBox.currentText()
        query_idx = list(QUERIES.keys()).index(query_name)
        params_amount = len([param for param in PARAMS[query_idx].split(" ") if param])

        self.params_textEdit.setText("")
        self.params_textEdit.setEnabled(bool(params_amount))
        self.params_textEdit.setPlaceholderText(PARAMS[query_idx])
        self.visualizeQuery_pushButton.setEnabled(not bool(params_amount))
        self.__clearTableWidget()

    def _findByCriteria(self) -> None:
        self.__clearTableWidget()
        self._visualizeTable()

    def _checkParamsAmount(self) -> None:
        query_name = self.queries_comboBox.currentText()
        query_idx = list(QUERIES.keys()).index(query_name)

        params_amount = len([param for param in PARAMS[query_idx].split(" ") if param])
        input_params = [input_param for input_param in self.params_textEdit.toPlainText().split("_") \
            if input_param]

        self.visualizeQuery_pushButton.setEnabled(len(input_params) == params_amount)

    @property
    def _selectedColumn(self) -> str:
        return REVERSED_COLUMNS_NAMES[self.columns_comboBox.currentText()]

    @property
    def _selectedTable(self) -> str:
        return TABLES_DICT[self.tables_comboBox.currentText()]

    @property
    def _criteriaPart(self) -> str:
        if len(self.columnCriteria_textEdit.toPlainText()):
            criteria_part = "WHERE {} = '{}' ".format(self._selectedColumn,\
                                                     self.columnCriteria_textEdit.toPlainText())
            return criteria_part
        
        return ""

    @property
    def _sortingPart(self) -> str:
        sorting_part = ""
        if self.enableSorting_checkBox.isChecked():
            sorting_part = "ORDER BY " + self._selectedColumn
            if self.sortByAscending_radioButton.isChecked():
                sorting_part += " ASC"
            else:
                sorting_part += " DESC"

        return sorting_part

    def __enableDMLButtons(self, state: bool) -> None:
        self.addRecord_pushButton.setEnabled(state)
        self.addRow_pushButton.setEnabled(state)
        self.deleteRecord_pushButton.setEnabled(state)
        self.updateRecord_pushButton.setEnabled(state)

    def _prepareCompoundForm(self) -> None:
        selected_table = TABLES_DICT[self.tables_comboBox.currentText()]
        self._cursor.execute("SELECT * FROM %s" % selected_table)

        main_table_columns_names: List[str] = [desc[0] for desc in self._cursor.description]
        main_table_columns_values: List[Any] = self._cursor.fetchall()

        child_table_names = CHILDREN_TABLES[self.tables_comboBox.currentText()]
        child_tables_columns_names: List[str] = []
        child_tables_columns_values: List[str] = []
        for child_table_name in child_table_names:
            self._cursor.execute("SELECT * FROM %s" % TABLES_DICT[child_table_name])
            child_tables_columns_names.append([desc[0] for desc in self._cursor.description])
            child_tables_columns_values.append(self._cursor.fetchall())

        self._openCompoundForm(selected_table,
                               main_table_columns_names, 
                               main_table_columns_values, 
                               child_table_names,
                               child_tables_columns_names, 
                               child_tables_columns_values)

    def _openCompoundForm(self,
                main_table_name: str,
                main_table_columns_names: List[str], 
                main_table_column_values: List[Any], 
                child_table_names: List[str],
                child_tables_columns_names: List[str], 
                child_tables_column_values: List[Any]) -> None:
        self.widget = CompoundForm(main_table_name,
                                   main_table_columns_names, 
                                   main_table_column_values, 
                                   child_table_names, 
                                   child_tables_columns_names, 
                                   child_tables_column_values,
                                   self._cursor,
                                   self._conn)
        self.widget.show()

    def _addRowToTheTableWidget(self) -> None:
        current_row_amount = self.tableWidget.rowCount()
        if current_row_amount == 0:
            return
        
        self.tableWidget.setRowCount(current_row_amount + 1)
        self._new_rows_added.append(current_row_amount + 1)

    def _addRecord(self) -> None:
        columns_names: List[str] = []
        if self._table_is_displayed:
            table_name = TABLES_DICT[self.tables_comboBox.currentText()]
        else:
            table_name = QUERIES[self.queries_comboBox.currentText()]
        for column in range(self.tableWidget.columnCount()):
            columns_names.append(self.tableWidget.horizontalHeaderItem(column).text())

        from datetime import datetime
        for row_idx in self._new_rows_added:
            row = []
            for column_idx in range(1, self.tableWidget.columnCount()):
                error_occured = False
                try:
                    _item_value = self.tableWidget.item(row_idx - 1, column_idx).text()
                    try:
                        _item_value = datetime.strptime(_item_value, "%Y-%m-%d")
                    except ValueError:
                        pass
                    row.append(_item_value)
                except AttributeError:
                    error_occured = True
                    break

            if error_occured:
                self.tableWidget.removeRow(row_idx)
                continue

            table_columns = ", ".join(columns_names[1:])
            table_values = ", ".join([f"'{value}'" for value in row])
            try:
                query = "INSERT INTO %s(%s) VALUES (%s);" % (table_name, table_columns, table_values)
                self._cursor.execute(query)
                self._conn.commit()
            except psycopg2.errors.InvalidDatetimeFormat:
                self._conn.rollback()
                return showError("Неверный формат даты\nПример правильного формата: 2001-01-30")

            self._default_rows_amount += 1
            return showMessage("Запись была успешно добавлена")

    def _addCellToArray(self) -> None:
        selected_item = self.tableWidget.selectedItems()[0]
        new_value = selected_item.text()

        selected_row = self.tableWidget.row(selected_item)
        if (selected_row + 1) > self._default_rows_amount:
            return
        
        seleted_column = self.tableWidget.column(selected_item)

        selected_column_name = self.tableWidget.horizontalHeaderItem(seleted_column).text()
        idx_column_value = self.tableWidget.item(selected_row, 0).text()
        
        if idx_column_value not in self._updatedRecordsInfo.keys():
            self._updatedRecordsInfo[idx_column_value] = {selected_column_name: new_value}
        else:
            self._updatedRecordsInfo[idx_column_value][selected_column_name] = new_value

    def _updateRecord(self) -> None:
        queries = []
        if self._table_is_displayed:
            table_name = TABLES_DICT[self.tables_comboBox.currentText()]
            main_column_name = self.tableWidget.horizontalHeaderItem(0).text()
        else:
            table_name = QUERIES[self.queries_comboBox.currentText()]
            main_column_name = self.tableWidget.horizontalHeaderItem(1).text()
        
        for main_column_idx, items_dict in self._updatedRecordsInfo.items():
            new_columns_values = []
            for column_name, new_value in items_dict.items():
                new_columns_values.append("{} = '{}'".format(column_name, new_value))

            if self._table_is_displayed:
                query = "UPDATE {} SET {} WHERE {} = {}".format(table_name, 
                            ", ".join(item for item in new_columns_values),
                            main_column_name,
                            main_column_idx)
            else:
                new_column_name = new_columns_values[0]
                procedures_names = [self.tableWidget.item(row, 1).text() for row in range(self.tableWidget.rowCount())]
                row_pos = procedures_names.index(new_column_name)

                old_column_value = self._default_procedures_names[row_pos]

                query = "UPDATE {} SET {} WHERE {} = {}".format(table_name, 
                new_column_name,
                main_column_name,
                old_column_value)
            queries.append(query)

        for query in queries:
            try:
                self._cursor.execute(query)
                self._conn.commit()
            except Exception as e:
                showError(str(e))
                self._conn.rollback()

        return showMessage("Все записи были успешно изменены")

    def _deleteRecord(self) -> None:
        current_table = TABLES_DICT[self.tables_comboBox.currentText()]
        selected_row = self.tableWidget.currentRow()
        unique_idx = self.tableWidget.item(selected_row, 0).text()
        
        self._cursor.execute("SELECT * FROM %s LIMIT 0" % (current_table, ))
        column_name = self._cursor.description[0][0]

        query = "DELETE FROM %s WHERE %s = %s" % (current_table, column_name, unique_idx)

        try:
            self._cursor.execute(query)
            self._conn.commit()
        except psycopg2.errors.InsufficientPrivilege:
            self._conn.rollback()
            return showError("У Вас недостаточно прав для удаленя данных")
        
        self.tableWidget.removeRow(selected_row)
        return showMessage("Запись была успешно удалена")

    def _visualizeTable(self) -> None:
        self.__enableDMLButtons(True)

        try:
            self._cursor.execute("SELECT * FROM %s %s %s" % (self._selectedTable, \
                self._criteriaPart, self._sortingPart))
        except psycopg2.errors.InsufficientPrivilege:
            showError("У Вас недостаточно прав для работы с данной таблицей")
            self._conn.rollback()
            return
        except psycopg2.errors.InvalidTextRepresentation:
            self._conn.rollback()
            return showError("Ошибка в формате данных")
        
        columns_names = [desc[0] for desc in self._cursor.description]
        self._table_is_displayed = True
        self.__fillData(columns_names)

    def _visualizeQuery(self) -> None:
        query_name = QUERIES[self.queries_comboBox.currentText()]

        input_params = [input_param for input_param in self.params_textEdit.toPlainText().split("_") \
            if input_param]
        input_params = ", ".join([f"'{input_param}'" for input_param in input_params])
        
        if not query_name == MODIFIED_VIEW:
            self.__enableDMLButtons(False)
        else:
            self.__enableDMLButtons(True)

        if "total" in query_name.lower():
            self.buildSummaryChart_pushButton.setEnabled(True)
            
        try:
            self._cursor.execute("SELECT * FROM %s" % (f"{query_name}({input_params})" if input_params else query_name))
        except psycopg2.errors.InsufficientPrivilege:
            self._conn.rollback()
            return showError("У Вас недостаточно прав для выполнения данного запроса")
        except psycopg2.errors.InvalidDatetimeFormat:
            self._conn.rollback()
            return showError("Неверный формат даты")
        except psycopg2.errors.InvalidTextRepresentation:
            self._conn.rollback()
            return showError("Неверный формат числа")
        
        columns_names = [desc[0] for desc in self._cursor.description]
        self._table_is_displayed = False
        self.__fillData(columns_names)

    def _updateSortingButtonsState(self) -> None:
        self.sortByAscending_radioButton.setEnabled(self.enableSorting_checkBox.isChecked())
        self.sortByDescending_radioButton.setEnabled(self.enableSorting_checkBox.isChecked())

    def __fillColumns(self) -> None:
        selected_table = TABLES_DICT[self.tables_comboBox.currentText()]
        try:
            self._cursor.execute("SELECT * FROM %s LIMIT 0;" % selected_table)
        except psycopg2.errors.InsufficientPrivilege:
            self._conn.rollback()
            self.columns_comboBox.clear()
            return showError("Недостаточно прав")
        columns_names: List[str] = [desc[0] for desc in self._cursor.description]
        
        self.columns_comboBox.clear()
        for column_name in columns_names:
            self.columns_comboBox.addItem(COLUMNS_NAMES[column_name])

    def __fillData(self, columns_names: List[str]) -> None:
        try:
            self.tableWidget.cellChanged.disconnect(self._addCellToArray)
        except TypeError:
            pass

        columns_amount = len(columns_names)
        print(columns_names)
        russian_columns_amount = [COLUMNS_NAMES[column] for column in columns_names]
        
        self.tableWidget.setColumnCount(columns_amount)
        self.tableWidget.setHorizontalHeaderLabels(russian_columns_amount)

        data = self._cursor.fetchall()
        self.tableWidget.setRowCount(len(data))
        self._default_rows_amount = len(data)
        for row_idx, row in enumerate(data):
            for column_idx, item in enumerate(row):
                if not self._table_is_displayed and \
                    QUERIES[self.queries_comboBox.currentText()] == MODIFIED_VIEW and \
                    column_idx == 1:
                    self._default_procedures_names.append(item)

                _item = QtWidgets.QTableWidgetItem(str(item))
                if (QUERIES[self.queries_comboBox.currentText()] == MODIFIED_VIEW and \
                    column_idx == 0 and \
                    not self._table_is_displayed) or \
                    (not self._table_is_displayed and \
                    not QUERIES[self.queries_comboBox.currentText()] == MODIFIED_VIEW):
                    _item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)
                self.tableWidget.setItem(row_idx, column_idx, _item)

        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.cellChanged.connect(self._addCellToArray)

    def _saveToExcel(self) -> None:
        if self.tableWidget.rowCount() == 0:
            return

        columns_values = [self.tableWidget.horizontalHeaderItem(column).text() \
            for column in range(self.tableWidget.columnCount())]
        values: List[List[Any]] = []
        
        for row in range(self.tableWidget.rowCount()):
            _inner_values: List[Any] = []
            for column in range(self.tableWidget.columnCount()):
                _inner_values.append(self.tableWidget.item(row, column).text())

            values.append(_inner_values)

        if self._table_is_displayed:
            saveDataToExcel(columns_values, values, self.tables_comboBox.currentText())
        else:
            saveDataToExcel(columns_values, values, QUERIES[self.queries_comboBox.currentText()])

    def _buildSummaryChart(self) -> None:
        values: List[str] = []
        labels: List[str] = [] 
        if self.tableWidget.columnCount() > 1:
            for row in range(self.tableWidget.rowCount()):
                labels.append(self.tableWidget.item(row, 0).text())
                try:
                    values.append(int(self.tableWidget.item(row, 1).text()))
                except ValueError:
                    values.append(float(self.tableWidget.item(row, 1).text()))
        else:
            for column in range(self.tableWidget.columnCount()):
                try:
                    labels.append(self.tableWidget.horizontalHeaderItem(column).text())
                    values.append(int(self.tableWidget.item(0, column).text()))
                except ValueError:
                    labels.pop()

        drawPieChart(values, labels, self.queries_comboBox.currentText())


if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    if TESTING_ENABLED:
        widget = MainWindow([], [])
    else:
        widget = MainWindow()
    widget.show()
    app.exec_()
